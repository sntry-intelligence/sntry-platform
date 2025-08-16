"""
Business Directory Export Service
Handles multi-format data export functionality
"""
import csv
import json
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.business_directory.models import Business
from app.business_directory.schemas import BusinessSearchFilters

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting business data in various formats"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def _apply_filters(self, query, filters: Optional[BusinessSearchFilters] = None):
        """Apply search filters to query"""
        if not filters:
            return query.filter(Business.is_active == True)
        
        # Apply basic filters
        if filters.query:
            search_term = f"%{filters.query}%"
            query = query.filter(
                Business.name.ilike(search_term) |
                Business.description.ilike(search_term) |
                Business.raw_address.ilike(search_term)
            )
        
        if filters.category:
            query = query.filter(Business.category.ilike(f"%{filters.category}%"))
        
        if filters.location:
            location_term = f"%{filters.location}%"
            query = query.filter(
                Business.raw_address.ilike(location_term) |
                Business.standardized_address.ilike(location_term)
            )
        
        # Apply spatial filters if provided
        if all([filters.latitude, filters.longitude, filters.radius]):
            from geoalchemy2.functions import ST_DWithin, ST_MakePoint, ST_SetSRID
            point = ST_SetSRID(ST_MakePoint(filters.longitude, filters.latitude), 4326)
            radius_meters = filters.radius * 1000
            query = query.filter(ST_DWithin(Business.geom, point, radius_meters))
        
        return query.filter(Business.is_active == filters.is_active)    de
f export_to_csv(self, filters: Optional[BusinessSearchFilters] = None) -> str:
        """Export businesses to CSV format"""
        try:
            query = self.db.query(Business)
            query = self._apply_filters(query, filters)
            businesses = query.all()
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            headers = [
                'ID', 'Name', 'Category', 'Raw Address', 'Standardized Address',
                'Phone Number', 'Email', 'Website', 'Description', 'Operating Hours',
                'Rating', 'Latitude', 'Longitude', 'Google Place ID',
                'Source URL', 'Last Scraped', 'Last Geocoded', 'Is Active',
                'Scrape Status', 'Geocode Status'
            ]
            writer.writerow(headers)
            
            # Write data rows
            for business in businesses:
                row = [
                    business.id,
                    business.name,
                    business.category or '',
                    business.raw_address,
                    business.standardized_address or '',
                    business.phone_number or '',
                    business.email or '',
                    business.website or '',
                    business.description or '',
                    business.operating_hours or '',
                    float(business.rating) if business.rating else '',
                    float(business.latitude) if business.latitude else '',
                    float(business.longitude) if business.longitude else '',
                    business.google_place_id or '',
                    business.source_url,
                    business.last_scraped_at.isoformat() if business.last_scraped_at else '',
                    business.last_geocoded_at.isoformat() if business.last_geocoded_at else '',
                    business.is_active,
                    business.scrape_status or '',
                    business.geocode_status or ''
                ]
                writer.writerow(row)
            
            logger.info(f"Exported {len(businesses)} businesses to CSV")
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise    def ex
port_to_excel(self, filters: Optional[BusinessSearchFilters] = None) -> bytes:
        """Export businesses to Excel format with multiple sheets"""
        try:
            query = self.db.query(Business)
            query = self._apply_filters(query, filters)
            businesses = query.all()
            
            wb = Workbook()
            
            # Main businesses sheet
            ws_main = wb.active
            ws_main.title = "Businesses"
            
            # Header styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Write headers
            headers = [
                'ID', 'Name', 'Category', 'Raw Address', 'Standardized Address',
                'Phone Number', 'Email', 'Website', 'Description', 'Operating Hours',
                'Rating', 'Latitude', 'Longitude', 'Google Place ID',
                'Source URL', 'Last Scraped', 'Last Geocoded', 'Is Active'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws_main.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Write data
            for row, business in enumerate(businesses, 2):
                ws_main.cell(row=row, column=1, value=business.id)
                ws_main.cell(row=row, column=2, value=business.name)
                ws_main.cell(row=row, column=3, value=business.category or '')
                ws_main.cell(row=row, column=4, value=business.raw_address)
                ws_main.cell(row=row, column=5, value=business.standardized_address or '')
                ws_main.cell(row=row, column=6, value=business.phone_number or '')
                ws_main.cell(row=row, column=7, value=business.email or '')
                ws_main.cell(row=row, column=8, value=business.website or '')
                ws_main.cell(row=row, column=9, value=business.description or '')
                ws_main.cell(row=row, column=10, value=business.operating_hours or '')
                ws_main.cell(row=row, column=11, value=float(business.rating) if business.rating else None)
                ws_main.cell(row=row, column=12, value=float(business.latitude) if business.latitude else None)
                ws_main.cell(row=row, column=13, value=float(business.longitude) if business.longitude else None)
                ws_main.cell(row=row, column=14, value=business.google_place_id or '')
                ws_main.cell(row=row, column=15, value=business.source_url)
                ws_main.cell(row=row, column=16, value=business.last_scraped_at)
                ws_main.cell(row=row, column=17, value=business.last_geocoded_at)
                ws_main.cell(row=row, column=18, value=business.is_active)
            
            # Auto-adjust column widths
            for column in ws_main.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_main.column_dimensions[column_letter].width = adjusted_width  
          # Summary sheet
            ws_summary = wb.create_sheet("Summary")
            ws_summary.cell(row=1, column=1, value="Export Summary").font = header_font
            ws_summary.cell(row=2, column=1, value="Total Businesses:")
            ws_summary.cell(row=2, column=2, value=len(businesses))
            ws_summary.cell(row=3, column=1, value="Export Date:")
            ws_summary.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            # Category breakdown
            categories = {}
            geocoded_count = 0
            for business in businesses:
                cat = business.category or "Uncategorized"
                categories[cat] = categories.get(cat, 0) + 1
                if business.latitude and business.longitude:
                    geocoded_count += 1
            
            ws_summary.cell(row=4, column=1, value="Geocoded Businesses:")
            ws_summary.cell(row=4, column=2, value=geocoded_count)
            ws_summary.cell(row=5, column=1, value="Geocoding Rate:")
            ws_summary.cell(row=5, column=2, value=f"{geocoded_count/len(businesses)*100:.1f}%" if businesses else "0%")
            
            # Categories sheet
            ws_categories = wb.create_sheet("Categories")
            ws_categories.cell(row=1, column=1, value="Category").font = header_font
            ws_categories.cell(row=1, column=2, value="Count").font = header_font
            
            for row, (category, count) in enumerate(sorted(categories.items(), key=lambda x: x[1], reverse=True), 2):
                ws_categories.cell(row=row, column=1, value=category)
                ws_categories.cell(row=row, column=2, value=count)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(f"Exported {len(businesses)} businesses to Excel with {len(categories)} categories")
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise    
def export_to_geojson(self, filters: Optional[BusinessSearchFilters] = None) -> Dict[str, Any]:
        """Export businesses to GeoJSON format for mapping"""
        try:
            query = self.db.query(Business)
            query = self._apply_filters(query, filters)
            # Only include businesses with coordinates for GeoJSON
            query = query.filter(Business.latitude.isnot(None), Business.longitude.isnot(None))
            businesses = query.all()
            
            features = []
            for business in businesses:
                feature = {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [float(business.longitude), float(business.latitude)]
                    },
                    "properties": {
                        "id": business.id,
                        "name": business.name,
                        "category": business.category,
                        "address": business.standardized_address or business.raw_address,
                        "phone": business.phone_number,
                        "email": business.email,
                        "website": business.website,
                        "description": business.description,
                        "rating": float(business.rating) if business.rating else None,
                        "operating_hours": business.operating_hours,
                        "google_place_id": business.google_place_id,
                        "last_updated": business.last_scraped_at.isoformat() if business.last_scraped_at else None
                    }
                }
                features.append(feature)
            
            geojson = {
                "type": "FeatureCollection",
                "features": features,
                "metadata": {
                    "total_features": len(features),
                    "export_date": datetime.now().isoformat(),
                    "filters_applied": filters.dict() if filters else None
                }
            }
            
            logger.info(f"Exported {len(features)} geocoded businesses to GeoJSON")
            return geojson
            
        except Exception as e:
            logger.error(f"Error exporting to GeoJSON: {e}")
            raise    de
f export_leads_to_csv(self, filters: Optional[BusinessSearchFilters] = None, 
                           min_lead_score: float = 0.0) -> str:
        """Export qualified leads to CSV format"""
        try:
            query = self.db.query(Business)
            query = self._apply_filters(query, filters)
            
            # Lead qualification criteria
            query = query.filter(
                Business.is_active == True,
                Business.latitude.isnot(None),  # Must be geocoded
                Business.longitude.isnot(None)
            )
            
            # Prefer businesses with contact information
            businesses = query.all()
            
            # Calculate lead scores and filter
            qualified_leads = []
            for business in businesses:
                lead_score = self._calculate_lead_score(business)
                if lead_score >= min_lead_score:
                    qualified_leads.append((business, lead_score))
            
            # Sort by lead score descending
            qualified_leads.sort(key=lambda x: x[1], reverse=True)
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            headers = [
                'Lead Score', 'Business Name', 'Category', 'Address', 'Phone', 'Email', 
                'Website', 'Rating', 'Description', 'Latitude', 'Longitude',
                'Contact Quality', 'Data Completeness', 'Last Updated'
            ]
            writer.writerow(headers)
            
            # Write lead data
            for business, score in qualified_leads:
                contact_quality = self._assess_contact_quality(business)
                data_completeness = self._calculate_data_completeness(business)
                
                row = [
                    f"{score:.2f}",
                    business.name,
                    business.category or '',
                    business.standardized_address or business.raw_address,
                    business.phone_number or '',
                    business.email or '',
                    business.website or '',
                    float(business.rating) if business.rating else '',
                    business.description or '',
                    float(business.latitude) if business.latitude else '',
                    float(business.longitude) if business.longitude else '',
                    contact_quality,
                    f"{data_completeness:.1f}%",
                    business.last_scraped_at.strftime("%Y-%m-%d") if business.last_scraped_at else ''
                ]
                writer.writerow(row)
            
            logger.info(f"Exported {len(qualified_leads)} qualified leads to CSV")
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting leads to CSV: {e}")
            raise    def _c
alculate_lead_score(self, business: Business) -> float:
        """Calculate lead score based on data quality and completeness"""
        score = 0.0
        
        # Base score for active business
        if business.is_active:
            score += 10.0
        
        # Geocoding bonus
        if business.latitude and business.longitude:
            score += 15.0
        
        # Contact information bonuses
        if business.phone_number:
            score += 20.0
        if business.email:
            score += 25.0
        if business.website:
            score += 15.0
        
        # Content quality bonuses
        if business.description and len(business.description) > 50:
            score += 10.0
        if business.category:
            score += 5.0
        if business.rating and business.rating >= 4.0:
            score += 10.0
        
        # Recency bonus
        if business.last_scraped_at:
            days_old = (datetime.now() - business.last_scraped_at).days
            if days_old <= 30:
                score += 5.0
            elif days_old <= 90:
                score += 2.0
        
        return min(score, 100.0)  # Cap at 100
    
    def _assess_contact_quality(self, business: Business) -> str:
        """Assess the quality of contact information"""
        contact_points = 0
        if business.phone_number:
            contact_points += 1
        if business.email:
            contact_points += 1
        if business.website:
            contact_points += 1
        
        if contact_points >= 3:
            return "Excellent"
        elif contact_points == 2:
            return "Good"
        elif contact_points == 1:
            return "Fair"
        else:
            return "Poor"
    
    def _calculate_data_completeness(self, business: Business) -> float:
        """Calculate percentage of data completeness"""
        total_fields = 12
        completed_fields = 0
        
        # Required fields
        if business.name:
            completed_fields += 1
        if business.raw_address:
            completed_fields += 1
        
        # Optional but valuable fields
        if business.category:
            completed_fields += 1
        if business.standardized_address:
            completed_fields += 1
        if business.phone_number:
            completed_fields += 1
        if business.email:
            completed_fields += 1
        if business.website:
            completed_fields += 1
        if business.description:
            completed_fields += 1
        if business.operating_hours:
            completed_fields += 1
        if business.rating:
            completed_fields += 1
        if business.latitude and business.longitude:
            completed_fields += 1
        if business.google_place_id:
            completed_fields += 1
        
        return (completed_fields / total_fields) * 100    de
f export_to_kml(self, filters: Optional[BusinessSearchFilters] = None) -> str:
        """Export businesses to KML format for Google MyMaps"""
        try:
            query = self.db.query(Business)
            query = self._apply_filters(query, filters)
            # Only include businesses with coordinates for KML
            query = query.filter(Business.latitude.isnot(None), Business.longitude.isnot(None))
            businesses = query.all()
            
            # KML header
            kml_content = [
                '<?xml version="1.0" encoding="UTF-8"?>',
                '<kml xmlns="http://www.opengis.net/kml/2.2">',
                '<Document>',
                '<name>Jamaica Business Directory</name>',
                '<description>Business locations exported from Jamaica Business Directory</description>',
                ''
            ]
            
            # Define styles for different business categories
            category_styles = self._generate_kml_styles()
            kml_content.extend(category_styles)
            
            # Group businesses by category for better organization
            businesses_by_category = {}
            for business in businesses:
                category = business.category or "Uncategorized"
                if category not in businesses_by_category:
                    businesses_by_category[category] = []
                businesses_by_category[category].append(business)
            
            # Create folders for each category
            for category, category_businesses in sorted(businesses_by_category.items()):
                kml_content.extend([
                    f'<Folder>',
                    f'<name>{self._escape_xml(category)} ({len(category_businesses)})</name>',
                    f'<description>Businesses in the {category} category</description>',
                    ''
                ])
                
                # Add placemarks for businesses in this category
                for business in category_businesses:
                    placemark = self._create_kml_placemark(business, category)
                    kml_content.extend(placemark)
                
                kml_content.extend(['</Folder>', ''])
            
            # KML footer
            kml_content.extend([
                '</Document>',
                '</kml>'
            ])
            
            kml_string = '\n'.join(kml_content)
            logger.info(f"Exported {len(businesses)} businesses to KML with {len(businesses_by_category)} categories")
            return kml_string
            
        except Exception as e:
            logger.error(f"Error exporting to KML: {e}")
            raise  
  def _generate_kml_styles(self) -> List[str]:
        """Generate KML styles for different business categories"""
        styles = []
        
        # Define category-specific styles with colors and icons
        category_styles = {
            "Restaurant": {"color": "ff0000ff", "icon": "http://maps.google.com/mapfiles/kml/pal2/icon63.png"},
            "Hotel": {"color": "ff00ff00", "icon": "http://maps.google.com/mapfiles/kml/pal2/icon62.png"},
            "Shopping": {"color": "ffff0000", "icon": "http://maps.google.com/mapfiles/kml/pal2/icon23.png"},
            "Healthcare": {"color": "ff00ffff", "icon": "http://maps.google.com/mapfiles/kml/pal2/icon61.png"},
            "Education": {"color": "ffff00ff", "icon": "http://maps.google.com/mapfiles/kml/pal2/icon58.png"},
            "Finance": {"color": "ff808080", "icon": "http://maps.google.com/mapfiles/kml/pal2/icon27.png"},
            "Automotive": {"color": "ff800080", "icon": "http://maps.google.com/mapfiles/kml/pal2/icon24.png"},
            "Entertainment": {"color": "ff008080", "icon": "http://maps.google.com/mapfiles/kml/pal2/icon26.png"},
            "Professional Services": {"color": "ff404040", "icon": "http://maps.google.com/mapfiles/kml/pal2/icon56.png"},
            "Uncategorized": {"color": "ff808080", "icon": "http://maps.google.com/mapfiles/kml/pal2/icon48.png"}
        }
        
        for category, style_info in category_styles.items():
            style_id = self._get_style_id(category)
            styles.extend([
                f'<Style id="{style_id}">',
                '<IconStyle>',
                f'<color>{style_info["color"]}</color>',
                '<scale>1.0</scale>',
                '<Icon>',
                f'<href>{style_info["icon"]}</href>',
                '</Icon>',
                '</IconStyle>',
                '<LabelStyle>',
                '<scale>0.8</scale>',
                '</LabelStyle>',
                '</Style>',
                ''
            ])
        
        return styles
    
    def _get_style_id(self, category: str) -> str:
        """Get style ID for a category"""
        # Convert category to valid XML ID
        style_id = category.lower().replace(" ", "_").replace("&", "and")
        return f"style_{style_id}"
    
    def _create_kml_placemark(self, business: Business, category: str) -> List[str]:
        """Create KML placemark for a business"""
        style_id = self._get_style_id(category)
        
        # Build description with business details
        description_parts = []
        
        if business.category:
            description_parts.append(f"<b>Category:</b> {self._escape_xml(business.category)}")
        
        address = business.standardized_address or business.raw_address
        description_parts.append(f"<b>Address:</b> {self._escape_xml(address)}")
        
        if business.phone_number:
            description_parts.append(f"<b>Phone:</b> {self._escape_xml(business.phone_number)}")
        
        if business.email:
            description_parts.append(f"<b>Email:</b> <a href='mailto:{business.email}'>{self._escape_xml(business.email)}</a>")
        
        if business.website:
            website_url = business.website if business.website.startswith('http') else f"http://{business.website}"
            description_parts.append(f"<b>Website:</b> <a href='{website_url}' target='_blank'>{self._escape_xml(business.website)}</a>")
        
        if business.rating:
            stars = "★" * int(float(business.rating)) + "☆" * (5 - int(float(business.rating)))
            description_parts.append(f"<b>Rating:</b> {stars} ({business.rating}/5)")
        
        if business.operating_hours:
            description_parts.append(f"<b>Hours:</b> {self._escape_xml(business.operating_hours)}")
        
        if business.description:
            # Truncate long descriptions
            desc_text = business.description[:200] + "..." if len(business.description) > 200 else business.description
            description_parts.append(f"<b>Description:</b> {self._escape_xml(desc_text)}")
        
        if business.last_scraped_at:
            description_parts.append(f"<b>Last Updated:</b> {business.last_scraped_at.strftime('%Y-%m-%d')}")
        
        description = "<br/>".join(description_parts)
        
        placemark = [
            '<Placemark>',
            f'<name>{self._escape_xml(business.name)}</name>',
            f'<description><![CDATA[{description}]]></description>',
            f'<styleUrl>#{style_id}</styleUrl>',
            '<Point>',
            f'<coordinates>{business.longitude},{business.latitude},0</coordinates>',
            '</Point>',
            '</Placemark>',
            ''
        ]
        
        return placemark
    
    def _escape_xml(self, text: str) -> str:
        """Escape XML special characters"""
        if not text:
            return ""
        
        text = str(text)
        text = text.replace("&", "&amp;")
        text = text.replace("<", "&lt;")
        text = text.replace(">", "&gt;")
        text = text.replace('"', "&quot;")
        text = text.replace("'", "&apos;")
        return text    def 
export_customer_360_to_excel(self, customer_ids: Optional[List[int]] = None) -> bytes:
        """Export comprehensive customer 360 profiles to Excel"""
        try:
            from app.customer_360.models import Customer, CustomerInteraction, CustomerBusinessRelationship
            
            # Get customers
            customer_query = self.db.query(Customer).filter(Customer.is_active == True)
            if customer_ids:
                customer_query = customer_query.filter(Customer.id.in_(customer_ids))
            customers = customer_query.all()
            
            wb = Workbook()
            
            # Customer profiles sheet
            ws_customers = wb.active
            ws_customers.title = "Customer Profiles"
            
            # Header styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Customer headers
            customer_headers = [
                'Customer ID', 'External ID', 'First Name', 'Last Name', 'Email', 'Phone',
                'Company', 'Industry', 'Customer Type', 'Lead Score', 'Lead Status',
                'Source System', 'Last Interaction', 'Total Interactions', 'Created Date'
            ]
            
            for col, header in enumerate(customer_headers, 1):
                cell = ws_customers.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Customer data
            for row, customer in enumerate(customers, 2):
                # Count interactions
                interaction_count = self.db.query(CustomerInteraction).filter(
                    CustomerInteraction.customer_id == customer.id
                ).count()
                
                ws_customers.cell(row=row, column=1, value=customer.id)
                ws_customers.cell(row=row, column=2, value=customer.external_id or '')
                ws_customers.cell(row=row, column=3, value=customer.first_name or '')
                ws_customers.cell(row=row, column=4, value=customer.last_name or '')
                ws_customers.cell(row=row, column=5, value=customer.email or '')
                ws_customers.cell(row=row, column=6, value=customer.phone_number or '')
                ws_customers.cell(row=row, column=7, value=customer.company_name or '')
                ws_customers.cell(row=row, column=8, value=customer.industry or '')
                ws_customers.cell(row=row, column=9, value=customer.customer_type or '')
                ws_customers.cell(row=row, column=10, value=float(customer.lead_score) if customer.lead_score else 0)
                ws_customers.cell(row=row, column=11, value=customer.lead_status or '')
                ws_customers.cell(row=row, column=12, value=customer.source_system or '')
                ws_customers.cell(row=row, column=13, value=customer.last_interaction_at)
                ws_customers.cell(row=row, column=14, value=interaction_count)
                ws_customers.cell(row=row, column=15, value=customer.created_at)
            
            # Auto-adjust column widths
            for column in ws_customers.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_customers.column_dimensions[column_letter].width = adjusted_width     
       # Interactions sheet
            ws_interactions = wb.create_sheet("Customer Interactions")
            
            interaction_headers = [
                'Interaction ID', 'Customer ID', 'Customer Name', 'Type', 'Channel',
                'Subject', 'Description', 'Outcome', 'Date', 'Duration (min)', 'Created By'
            ]
            
            for col, header in enumerate(interaction_headers, 1):
                cell = ws_interactions.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Get all interactions for the customers
            interaction_query = self.db.query(CustomerInteraction).join(Customer)
            if customer_ids:
                interaction_query = interaction_query.filter(Customer.id.in_(customer_ids))
            interactions = interaction_query.all()
            
            for row, interaction in enumerate(interactions, 2):
                customer_name = f"{interaction.customer.first_name or ''} {interaction.customer.last_name or ''}".strip()
                
                ws_interactions.cell(row=row, column=1, value=interaction.id)
                ws_interactions.cell(row=row, column=2, value=interaction.customer_id)
                ws_interactions.cell(row=row, column=3, value=customer_name)
                ws_interactions.cell(row=row, column=4, value=interaction.interaction_type)
                ws_interactions.cell(row=row, column=5, value=interaction.interaction_channel or '')
                ws_interactions.cell(row=row, column=6, value=interaction.subject or '')
                ws_interactions.cell(row=row, column=7, value=interaction.description or '')
                ws_interactions.cell(row=row, column=8, value=interaction.outcome or '')
                ws_interactions.cell(row=row, column=9, value=interaction.interaction_date)
                ws_interactions.cell(row=row, column=10, value=interaction.duration_minutes or '')
                ws_interactions.cell(row=row, column=11, value=interaction.created_by or '')
            
            # Business relationships sheet
            ws_relationships = wb.create_sheet("Business Relationships")
            
            relationship_headers = [
                'Relationship ID', 'Customer ID', 'Customer Name', 'Business ID', 'Business Name',
                'Relationship Type', 'Status', 'Start Date', 'End Date', 'Notes'
            ]
            
            for col, header in enumerate(relationship_headers, 1):
                cell = ws_relationships.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Get business relationships
            relationship_query = self.db.query(CustomerBusinessRelationship).join(Customer)
            if customer_ids:
                relationship_query = relationship_query.filter(Customer.id.in_(customer_ids))
            relationships = relationship_query.all()
            
            for row, rel in enumerate(relationships, 2):
                customer_name = f"{rel.customer.first_name or ''} {rel.customer.last_name or ''}".strip()
                
                # Get business name
                business = self.db.query(Business).filter(Business.id == rel.business_id).first()
                business_name = business.name if business else f"Business ID {rel.business_id}"
                
                ws_relationships.cell(row=row, column=1, value=rel.id)
                ws_relationships.cell(row=row, column=2, value=rel.customer_id)
                ws_relationships.cell(row=row, column=3, value=customer_name)
                ws_relationships.cell(row=row, column=4, value=rel.business_id)
                ws_relationships.cell(row=row, column=5, value=business_name)
                ws_relationships.cell(row=row, column=6, value=rel.relationship_type)
                ws_relationships.cell(row=row, column=7, value=rel.relationship_status)
                ws_relationships.cell(row=row, column=8, value=rel.start_date)
                ws_relationships.cell(row=row, column=9, value=rel.end_date)
                ws_relationships.cell(row=row, column=10, value=rel.notes or '')
            
            # Summary sheet
            ws_summary = wb.create_sheet("Summary")
            ws_summary.cell(row=1, column=1, value="Customer 360 Export Summary").font = header_font
            ws_summary.cell(row=2, column=1, value="Total Customers:")
            ws_summary.cell(row=2, column=2, value=len(customers))
            ws_summary.cell(row=3, column=1, value="Total Interactions:")
            ws_summary.cell(row=3, column=2, value=len(interactions))
            ws_summary.cell(row=4, column=1, value="Total Relationships:")
            ws_summary.cell(row=4, column=2, value=len(relationships))
            ws_summary.cell(row=5, column=1, value="Export Date:")
            ws_summary.cell(row=5, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            # Lead score distribution
            high_score_customers = len([c for c in customers if c.lead_score and float(c.lead_score) >= 75])
            medium_score_customers = len([c for c in customers if c.lead_score and 50 <= float(c.lead_score) < 75])
            low_score_customers = len([c for c in customers if c.lead_score and float(c.lead_score) < 50])
            
            ws_summary.cell(row=7, column=1, value="Lead Score Distribution:").font = header_font
            ws_summary.cell(row=8, column=1, value="High Score (75+):")
            ws_summary.cell(row=8, column=2, value=high_score_customers)
            ws_summary.cell(row=9, column=1, value="Medium Score (50-74):")
            ws_summary.cell(row=9, column=2, value=medium_score_customers)
            ws_summary.cell(row=10, column=1, value="Low Score (<50):")
            ws_summary.cell(row=10, column=2, value=low_score_customers)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(f"Exported {len(customers)} customer 360 profiles to Excel")
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting customer 360 to Excel: {e}")
            raise 
   def export_qualified_leads_to_csv(self, min_score: float = 70.0, 
                                     max_results: int = 1000) -> str:
        """Export sales-ready qualified leads to CSV"""
        try:
            from app.customer_360.models import Customer, CustomerInteraction, CustomerBusinessRelationship
            
            # Get qualified leads
            customer_query = self.db.query(Customer).filter(
                Customer.is_active == True,
                Customer.lead_score >= min_score
            ).order_by(Customer.lead_score.desc()).limit(max_results)
            
            customers = customer_query.all()
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            headers = [
                'Customer ID', 'Lead Score', 'Lead Status', 'First Name', 'Last Name',
                'Email', 'Phone', 'Company', 'Industry', 'Customer Type',
                'Last Interaction', 'Total Interactions', 'Business Connections',
                'Qualification Reasons', 'Recommended Actions', 'Territory',
                'Source System', 'Created Date'
            ]
            writer.writerow(headers)
            
            # Process each customer
            for customer in customers:
                # Get interaction count
                interaction_count = self.db.query(CustomerInteraction).filter(
                    CustomerInteraction.customer_id == customer.id
                ).count()
                
                # Get business connections
                business_connections = self.db.query(CustomerBusinessRelationship).filter(
                    CustomerBusinessRelationship.customer_id == customer.id,
                    CustomerBusinessRelationship.relationship_status == 'active'
                ).count()
                
                # Generate qualification reasons
                qualification_reasons = self._generate_qualification_reasons(customer, interaction_count)
                
                # Generate recommended actions
                recommended_actions = self._generate_recommended_actions(customer, interaction_count)
                
                # Determine territory (simplified - could be enhanced with geographic data)
                territory = self._determine_territory(customer)
                
                row = [
                    customer.id,
                    float(customer.lead_score) if customer.lead_score else 0,
                    customer.lead_status or '',
                    customer.first_name or '',
                    customer.last_name or '',
                    customer.email or '',
                    customer.phone_number or '',
                    customer.company_name or '',
                    customer.industry or '',
                    customer.customer_type or '',
                    customer.last_interaction_at.strftime("%Y-%m-%d") if customer.last_interaction_at else '',
                    interaction_count,
                    business_connections,
                    '; '.join(qualification_reasons),
                    '; '.join(recommended_actions),
                    territory,
                    customer.source_system or '',
                    customer.created_at.strftime("%Y-%m-%d") if customer.created_at else ''
                ]
                writer.writerow(row)
            
            logger.info(f"Exported {len(customers)} qualified leads to CSV")
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting qualified leads to CSV: {e}")
            raise    d
ef export_dashboard_data_to_json(self) -> Dict[str, Any]:
        """Export analytics dashboard data for BI tool integration"""
        try:
            from app.customer_360.models import Customer, CustomerInteraction, CustomerBusinessRelationship
            
            # Business metrics
            total_businesses = self.db.query(Business).filter(Business.is_active == True).count()
            geocoded_businesses = self.db.query(Business).filter(
                Business.is_active == True,
                Business.latitude.isnot(None),
                Business.longitude.isnot(None)
            ).count()
            
            # Customer metrics
            total_customers = self.db.query(Customer).filter(Customer.is_active == True).count()
            qualified_leads = self.db.query(Customer).filter(
                Customer.is_active == True,
                Customer.lead_score >= 70
            ).count()
            
            # Interaction metrics
            total_interactions = self.db.query(CustomerInteraction).count()
            recent_interactions = self.db.query(CustomerInteraction).filter(
                CustomerInteraction.interaction_date >= datetime.now().replace(day=1)  # This month
            ).count()
            
            # Category distribution
            category_query = self.db.query(
                Business.category,
                func.count(Business.id).label('count')
            ).filter(Business.is_active == True).group_by(Business.category)
            
            categories = {}
            for category, count in category_query.all():
                categories[category or 'Uncategorized'] = count
            
            # Lead score distribution
            lead_score_ranges = {
                'high_score': self.db.query(Customer).filter(
                    Customer.is_active == True,
                    Customer.lead_score >= 75
                ).count(),
                'medium_score': self.db.query(Customer).filter(
                    Customer.is_active == True,
                    Customer.lead_score >= 50,
                    Customer.lead_score < 75
                ).count(),
                'low_score': self.db.query(Customer).filter(
                    Customer.is_active == True,
                    Customer.lead_score < 50
                ).count()
            }
            
            # Industry distribution
            industry_query = self.db.query(
                Customer.industry,
                func.count(Customer.id).label('count')
            ).filter(Customer.is_active == True).group_by(Customer.industry)
            
            industries = {}
            for industry, count in industry_query.all():
                industries[industry or 'Unknown'] = count
            
            # Monthly trends (last 12 months)
            monthly_trends = self._calculate_monthly_trends()
            
            dashboard_data = {
                'export_metadata': {
                    'export_date': datetime.now().isoformat(),
                    'data_freshness': 'real-time',
                    'version': '1.0'
                },
                'business_metrics': {
                    'total_businesses': total_businesses,
                    'geocoded_businesses': geocoded_businesses,
                    'geocoding_rate': (geocoded_businesses / total_businesses * 100) if total_businesses > 0 else 0,
                    'category_distribution': categories
                },
                'customer_metrics': {
                    'total_customers': total_customers,
                    'qualified_leads': qualified_leads,
                    'qualification_rate': (qualified_leads / total_customers * 100) if total_customers > 0 else 0,
                    'lead_score_distribution': lead_score_ranges,
                    'industry_distribution': industries
                },
                'interaction_metrics': {
                    'total_interactions': total_interactions,
                    'recent_interactions': recent_interactions,
                    'avg_interactions_per_customer': (total_interactions / total_customers) if total_customers > 0 else 0
                },
                'trends': monthly_trends,
                'kpis': {
                    'data_quality_score': self._calculate_data_quality_score(),
                    'lead_conversion_rate': self._calculate_lead_conversion_rate(),
                    'customer_engagement_score': self._calculate_engagement_score()
                }
            }
            
            logger.info("Exported dashboard data for BI integration")
            return dashboard_data
            
        except Exception as e:
            logger.error(f"Error exporting dashboard data: {e}")
            raise    def 
_generate_qualification_reasons(self, customer, interaction_count: int) -> List[str]:
        """Generate reasons why a customer is qualified"""
        reasons = []
        
        if customer.lead_score and float(customer.lead_score) >= 90:
            reasons.append("Exceptional lead score")
        elif customer.lead_score and float(customer.lead_score) >= 75:
            reasons.append("High lead score")
        
        if interaction_count >= 5:
            reasons.append("High engagement level")
        elif interaction_count >= 2:
            reasons.append("Active engagement")
        
        if customer.company_name:
            reasons.append("Business customer")
        
        if customer.email and customer.phone_number:
            reasons.append("Complete contact information")
        
        if customer.industry:
            reasons.append(f"Targeted industry: {customer.industry}")
        
        if customer.last_interaction_at:
            days_since_interaction = (datetime.now() - customer.last_interaction_at).days
            if days_since_interaction <= 30:
                reasons.append("Recent interaction")
        
        return reasons or ["Meets minimum qualification criteria"]
    
    def _generate_recommended_actions(self, customer, interaction_count: int) -> List[str]:
        """Generate recommended next actions for a lead"""
        actions = []
        
        if not customer.phone_number:
            actions.append("Obtain phone number")
        
        if not customer.email:
            actions.append("Collect email address")
        
        if interaction_count == 0:
            actions.append("Initial contact")
        elif interaction_count < 3:
            actions.append("Follow-up contact")
        else:
            actions.append("Schedule meeting")
        
        if customer.company_name and not customer.industry:
            actions.append("Identify industry sector")
        
        if customer.lead_status == 'new':
            actions.append("Qualify lead")
        elif customer.lead_status == 'qualified':
            actions.append("Present solution")
        
        if customer.last_interaction_at:
            days_since_interaction = (datetime.now() - customer.last_interaction_at).days
            if days_since_interaction > 30:
                actions.append("Re-engage customer")
        
        return actions or ["Contact customer"]
    
    def _determine_territory(self, customer) -> str:
        """Determine sales territory for customer"""
        # Simplified territory assignment - could be enhanced with geographic data
        if customer.address:
            address_lower = customer.address.lower()
            if 'kingston' in address_lower:
                return "Kingston Metro"
            elif 'spanish town' in address_lower:
                return "St. Catherine"
            elif 'montego bay' in address_lower:
                return "St. James"
            elif 'mandeville' in address_lower:
                return "Manchester"
            else:
                return "Other Parishes"
        
        return "Unassigned"    
def _calculate_monthly_trends(self) -> Dict[str, Any]:
        """Calculate monthly trends for the last 12 months"""
        try:
            from app.customer_360.models import Customer, CustomerInteraction
            from dateutil.relativedelta import relativedelta
            
            trends = {
                'new_customers': [],
                'new_businesses': [],
                'interactions': [],
                'months': []
            }
            
            # Calculate for last 12 months
            for i in range(12):
                month_start = datetime.now().replace(day=1) - relativedelta(months=i)
                month_end = month_start + relativedelta(months=1)
                
                # New customers
                new_customers = self.db.query(Customer).filter(
                    Customer.created_at >= month_start,
                    Customer.created_at < month_end
                ).count()
                
                # New businesses
                new_businesses = self.db.query(Business).filter(
                    Business.last_scraped_at >= month_start,
                    Business.last_scraped_at < month_end
                ).count()
                
                # Interactions
                interactions = self.db.query(CustomerInteraction).filter(
                    CustomerInteraction.interaction_date >= month_start,
                    CustomerInteraction.interaction_date < month_end
                ).count()
                
                trends['new_customers'].insert(0, new_customers)
                trends['new_businesses'].insert(0, new_businesses)
                trends['interactions'].insert(0, interactions)
                trends['months'].insert(0, month_start.strftime('%Y-%m'))
            
            return trends
            
        except Exception as e:
            logger.error(f"Error calculating monthly trends: {e}")
            return {'error': str(e)}
    
    def _calculate_data_quality_score(self) -> float:
        """Calculate overall data quality score"""
        try:
            total_businesses = self.db.query(Business).filter(Business.is_active == True).count()
            if total_businesses == 0:
                return 0.0
            
            # Count businesses with complete data
            complete_businesses = self.db.query(Business).filter(
                Business.is_active == True,
                Business.name.isnot(None),
                Business.raw_address.isnot(None),
                Business.latitude.isnot(None),
                Business.longitude.isnot(None)
            ).count()
            
            return (complete_businesses / total_businesses) * 100
            
        except Exception as e:
            logger.error(f"Error calculating data quality score: {e}")
            return 0.0
    
    def _calculate_lead_conversion_rate(self) -> float:
        """Calculate lead conversion rate"""
        try:
            from app.customer_360.models import Customer
            
            total_leads = self.db.query(Customer).filter(
                Customer.is_active == True,
                Customer.lead_score >= 50
            ).count()
            
            if total_leads == 0:
                return 0.0
            
            converted_leads = self.db.query(Customer).filter(
                Customer.is_active == True,
                Customer.lead_status == 'converted'
            ).count()
            
            return (converted_leads / total_leads) * 100
            
        except Exception as e:
            logger.error(f"Error calculating lead conversion rate: {e}")
            return 0.0
    
    def _calculate_engagement_score(self) -> float:
        """Calculate customer engagement score"""
        try:
            from app.customer_360.models import Customer, CustomerInteraction
            
            total_customers = self.db.query(Customer).filter(Customer.is_active == True).count()
            if total_customers == 0:
                return 0.0
            
            # Customers with recent interactions (last 90 days)
            recent_date = datetime.now() - relativedelta(days=90)
            engaged_customers = self.db.query(Customer).filter(
                Customer.is_active == True,
                Customer.last_interaction_at >= recent_date
            ).count()
            
            return (engaged_customers / total_customers) * 100
            
        except Exception as e:
            logger.error(f"Error calculating engagement score: {e}")
            return 0.0